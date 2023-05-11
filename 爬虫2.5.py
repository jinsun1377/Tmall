import requests
import json
import time
import re
import time
from loguru import logger
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from bs4 import BeautifulSoup
import urllib3
import urllib.parse


# 去除网页验证警告
urllib3.disable_warnings()


class Tmall():
    def __init__(self):
        self.ti = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))

        self.sort_list = [
            "default",
            "sale-desc"
        ]
        with open('config.json', 'r', encoding='utf-8') as f:
            config_text = f.read()
            # print(config_text)
            if config_text.startswith(u'\ufeff'):
                config_text = config_text.encode('utf8')[3:].decode('utf8')
        try:
            self.search_info = json.loads(config_text)
        except Exception as e:
            logger.debug(e)
            logger.info('【config】配置文件不是json格式，请检查后重启')
            time.sleep(60 * 5)
        # 一个字典：id作为键，keyword作为值
        self.goods_id_keyword_dict = {}
        # 创建一个字典容纳所需要的的所有信息：商品id-店名-价格-付款人数-收货人数
        self.goods_info_list = {}

        # 将每日爬取的数据存放到文件夹‘淘宝数据’中，先判定下这个文件夹是否存在
        a = os.path.exists('淘宝数据')
        if not a:
            os.makedirs('淘宝数据')
        # 创建xlxs文件存放爬取到的数据
        self.wb = Workbook()
        # 获取当日日期作为文件名和工作簿名，注意加后缀‘.xlsx’，再加上相对路径
        self.localtime = time.strftime('%Y-%m-%d', time.localtime(time.time()))
        self.name = self.localtime + '.xlsx'
        self.path = './淘宝数据/' + self.name
        # 创建存储数据的xlsx表，以日期命名，并将表头设置好
        wb_sheet = self.wb.create_sheet(self.name, index=0)
        wb_sheet.cell(1, 1, '商品ID')
        wb_sheet.cell(1, 2, '商品名')
        wb_sheet.cell(1, 3, '店名')
        wb_sheet.cell(1, 4, '价格')
        wb_sheet.cell(1, 5, '付款人数')
        wb_sheet.cell(1, 6, '收货人数')
        # 将文件保存在文件夹‘淘宝数据’下面，相对路径
        self.wb.save(self.path)
        self.row = 2
        # 记录爬虫进程，以关键字为标准
        self.length_keyword = 1
        self.progress = self.length_keyword / len(self.search_info['info_list'])
        # 代理ip
        self.proxies = {'http': 'http://183.166.162.134'}

    # 在淘宝中爬取
    def get_taobao_search_list(self, keyword, sort):

        expection_count = 0
        while 1:
            try:
                headers = {
                    'pragma': 'no-cache',
                    'cache-control': 'no-cache',
                    'upgrade-insecure-requests': '1',
                    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
                    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
                    'accept-encoding': 'gzip, deflate, br',
                    'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8',
                    'cookie': self.search_info['Cookie'],
                }

                url = 'https://s.taobao.com/search?q={}&imgfile=&ie=utf8&sort={}'.format(urllib.parse.quote(keyword), sort)
                response = requests.get(url, headers=headers, verify=False)

                if "g_page_config " in response.text:
                    text = re.findall(r"g_page_config = (.*?)}};", response.text, re.S)[0]
                    json_data = json.loads(text + "}}")
                    return json_data
                else:
                    logger.info('提示跳转登陆页，请更换cookie--暂停5分钟')
                    logger.info('更换成功后，请重启本程序--暂停中')
                    time.sleep(60 * self.search_info['time'])
            except Exception as e:
                if expection_count > 5:
                    logger.debug(e)
                    logger.info('【get_taobao_search_list】出现错误，请联系作者')
                    time.sleep(60 * self.search_info['time'])
                else:
                    expection_count += 1
                    time.sleep(5)

    # 在天猫中爬取
    def get_Tmall_html(self, goods_id):
        expection_count = 0
        while 1:
            try:
                headers = {

                    'upgrade-insecure-requests': '1',
                    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
                    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
                    'accept-encoding': 'gzip, deflate, br',
                    'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8',
                }

                url = 'https://h5api.m.taobao.com/h5/mtop.taobao.detail.getdetail/6.0/?api={}&v=6.0&dataType=jsonp' \
                      '&ttid={}&AntiCreep=true&type=json&callback=mtopjsonp2&data=%7B%22itemNumId%22%3A%22{}%22%7D'.format(
                    'mtop.taobao.detail.getdetail', '2017@taobao_h5_6.6.0', goods_id
                )
                print(0)
                response = requests.get(url, headers=headers, verify=False)
                text = re.findall(r"mtopjsonp2\((.*?)}\)", response.text, re.S)[0]
                data1 = json.loads(text + "}")
                print(type(data1))
                print(data1["ret"])
                print(json.loads(data1['data']['apiStack'][0]['value']))
                return data1['ret'], json.loads(data1['data']['apiStack'][0]['value'])
            except Exception as e:

                if expection_count > 5:
                    logger.debug(e)
                    logger.info('【get_Tmall_html】出现错误，请联系作者')
                    time.sleep(60 * self.search_info['time'])
                else:
                    expection_count += 1
                    time.sleep(5)

    def run(self, good_id, keyword):
        # keyword = search_infos['search_keyword']# 关键字
        # print("搜索进度:{:.2%}".format(self.progress))
        print("搜索进度:" + str(self.length_keyword) + "/" + str(len(self.search_info['info_list'])))
        print("搜索商品ID:{}".format(good_id))
        self.goods_info_list[good_id] = {}
        #for goods_id in self.goods_id_list:
            # 将商品id编码作为键，值全为空
            #self.goods_info_list[keyword][goods_id] = {}
        print(self.goods_id_keyword_dict)
        print(keyword)
        for sort in self.sort_list:
            search_json = self.get_taobao_search_list(keyword, sort)
            # print(search_json)
            for item in search_json['mods']['itemlist']['data']['auctions']:
                price = item['view_price']
                sales = item['view_sales']
                shop_name = item['nick']
                # 一个关键字可能搜索出好几个店铺的信息
                goods_id = item['nid']
                # print(goods_id)
                # 同一本书两家店铺的关键字可能一模一样，所以仍需要判定一下
                # print(type(self.search_info['info_list'][0]))
                if goods_id in self.search_info['info_list']:
                    # print(item)
                    self.goods_info_list[good_id]['price'] = price
                    self.goods_info_list[good_id]['shop_name'] = shop_name
                    if sort == 'default':
                        self.goods_info_list[good_id]['payCount'] = sales
                    else:
                        self.goods_info_list[good_id]['takeCount'] = sales
        # 此时是根据id一个一个搜索，所以不再需要使用for循环，直接将值存入字典,并放入excel表格即可
        wb1 = load_workbook(self.path)
        sheet = wb1[self.name]
        sheet.cell(self.row, 1, good_id)
        sheet.cell(self.row, 2, keyword)
        sheet.cell(self.row, 3, self.goods_info_list[good_id].get("shop_name", "无"))
        sheet.cell(self.row, 4, self.goods_info_list[good_id].get("price", "无"))
        # 提取付款人数、收货人数中的数字提取出来
        sheet.cell(self.row, 5, re.sub('\D', '', self.goods_info_list[good_id].get("payCount", "无")))
        sheet.cell(self.row, 6, re.sub('\D', '', self.goods_info_list[good_id].get("takeCount", "无")))
        wb1.save(self.path)
        self.row += 1

        # 优化，是否可以存放完一个关键字相关信息便存储，这样即使程序意外中断，也可以获取部分信息
        #for good_id in self.goods_info_list[keyword]:
            #wb1 = load_workbook(self.path)
            #sheet = wb1[self.name]
            #sheet.cell(self.row, 1, keyword)
            #sheet.cell(self.row, 2, good_id)
            #sheet.cell(self.row, 3, self.goods_info_list[keyword][good_id].get("shop_name", "无"))
            #sheet.cell(self.row, 4, self.goods_info_list[keyword][good_id].get("price", "无"))
            # 提取付款人数、收货人数中的数字提取出来
            #sheet.cell(self.row, 5, re.sub('\D', '', self.goods_info_list[keyword][good_id].get("payCount", "无")))
            #sheet.cell(self.row, 6, re.sub('\D', '', self.goods_info_list[keyword][good_id].get("takeCount", "无")))
            #wb1.save(self.path)
            #self.row += 1


            #with open("result", "a") as f:
                #f.write('关键字:{}\n商品ID:{}\n店名:{}\n价格:{}\n付款人数:{}\n收货人数:{}\n'.format(
                    #keyword,
                    #good_id,
                    #self.goods_info_list[keyword][good_id].get("shop_name", "无"),
                    #self.goods_info_list[keyword][good_id].get("price", "无"),
                    #self.goods_info_list[keyword][good_id].get("payCount", "无"),
                    #self.goods_info_list[keyword][good_id].get("takeCount", "无")
                #))


        #for goods_id in self.goods_id_list:
            #call_result, data_json = self.get_Tmall_html(goods_id)
            #goods_info_list[goods_id]['monthSalesCount'] = data_json['item']['sellCount']

        # print(goods_info_list)
        #for goods_id in self.goods_id_list:
            #with open("result","a") as f:
                #f.write('--店名：【{}】--\n价格：{}\n--月销售：{}\n--付款人数：{}\n--收货人数：{}--\n'.format(
                    #goods_info_list[goods_id].get('shop_name', '无'),
                    #goods_info_list[goods_id].get('price', '【无】'),
                    #goods_info_list[goods_id].get('monthSalesCount', '【无】'),
                    #goods_info_list[goods_id].get('payCount', '【无】'),
                    #goods_info_list[goods_id].get('takeCount', '【无】'),

                #))

    # 获取id对应的关键字keyword，并放入字典中
    def get_keyword(self, good_id):
        expection_count = 0

        try:
            headers = {
                'pragma': 'no-cache',
                'cache-control': 'no-cache',
                'upgrade-insecure-requests': '1',
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
                'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
                'accept-encoding': 'gzip, deflate, br',
                'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8',
                # 不使用cookie爬取关键字，减少cookie被封的风险
                # 'cookie': self.search_info['Cookie'],
            }

            url = 'https://detail.tmall.com/item.htm?spm=a230r.1.14.16.7488a121Voc9YO&id={}&ns=1&abbucket=15'.format(good_id)
            response = requests.get(url, headers=headers, verify=False)
            soup = BeautifulSoup(response.text, 'lxml')
            print(response.text)
            keyword = soup.find('meta', {'name': 'keywords'})['content']
            # print(type(keyword))
            if keyword == "":
                logger.info('提示跳转登陆页，请更换cookie--暂停5分钟')
                logger.info('更换成功后，请重启本程序--暂停中')
                time.sleep(60 * self.search_info['time'])
            else:
                self.goods_id_keyword_dict[good_id] = keyword
            print(self.goods_id_keyword_dict)
        except Exception as e:
            if expection_count > 5:
                logger.debug(e)
                logger.info('【获得{}的关键字】失败，请联系作者'.format(good_id))
                time.sleep(60 * self.search_info['time'])
            else:
                expection_count += 1
                time.sleep(5)



    def main(self):

        while 1:
            print('--当前时间--【{}】--间隔时间：【{}】分钟--'.format(self.ti, self.search_info['time']))
            print('--本查询只支持【天猫】店铺查询--')

            try:
                for i in self.search_info['info_list']:
                    # for goods_url in i['goods_list']:
                    # 将所有要搜索商品的id加入goods_id_list
                    self.goods_id_keyword_dict[i] = ""
                    self.get_keyword(i)
                    # 将本次循环获得的id和keyword传入run函数
                    self.run(i, self.goods_id_keyword_dict[i])
                    # 更新搜索进度
                    self.length_keyword += 1
                    self.progress = self.length_keyword / len(self.search_info['info_list'])
                # 当把所有需要的信息都存放入self.goods_info_list字典后，进行提取

                #for keyword in self.goods_info_list:
                    #for good_id in self.goods_info_list[keyword]:
                        #with open("result", "a") as f:
                            #f.write('关键字:{}\n商品ID:{}\n店名:{}\n价格:{}\n付款人数:{}\n收货人数:{}\n'.format(
                                #keyword,
                                #good_id,
                                #self.goods_info_list[keyword][good_id].get("shop_name", "无"),
                                #self.goods_info_list[keyword][good_id].get("price", "无"),
                                #self.goods_info_list[keyword][good_id].get("payCount", "无"),
                                #self.goods_info_list[keyword][good_id].get("takeCount", "无")
                            #))
                    # 每次存储完id和对应的关键字后清空后重新存储
                    self.goods_id_keyword_dict.clear()
                print('=======本次搜索完毕，进入暂停=======')
                print('\n')
                time.sleep(self.search_info['time'] * 60)
            except Exception as e:
                logger.debug(e)
                logger.info('【config】中json文件读取出现问题，请检查后重启')
                time.sleep(60 * 5)


if __name__ == '__main__':
    tm = Tmall()
    tm.main()
